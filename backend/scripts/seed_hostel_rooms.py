import os
import re
import openpyxl
from pymongo import MongoClient

env_path = r"F:\Aarambh 2026\Team_portal\backend\.env"
excel_path = r"C:\Users\MSI1\Downloads\Hostel Room.xlsx"

def get_mongo_uri():
    with open(env_path, 'r') as f:
        content = f.read()
    match = re.search(r'MONGODB_URI\s*=\s*["\']?(.*?)["\']?$', content, re.MULTILINE)
    if match:
        return match.group(1).strip()
    return None

def parse_bh1(sheet):
    rooms = []
    current_floor = None
    current_room = None
    
    rows = list(sheet.iter_rows(values_only=True))[3:] # skip header lines
    
    for r in rows:
        floor = r[0]
        room_num = r[1]
        sno = r[2]
        name = r[3]
        committee = r[4]
        post = r[5]
        
        if floor: current_floor = str(floor).strip()
        if room_num: current_room = str(room_num).strip()
        
        if not current_room or sno is None:
            continue
            
        try:
            sno_val = int(sno)
        except ValueError:
            continue
            
        bed_num = f"Bed {((sno_val - 1) % 2) + 1}"
        
        rooms.append({
            'hostel': 'BH-1',
            'floor': current_floor,
            'room': current_room,
            'bed': bed_num,
            'sno': sno_val,
            'allottedTo': None,
            'allottedToName': None,
            'allottedToAppNo': None
        })
    return rooms

def parse_gh2(sheet):
    rooms = []
    current_floor = None
    current_room = None
    current_type = None
    
    rows = list(sheet.iter_rows(values_only=True))[3:] # skip header lines
    
    for r in rows:
        floor = r[0]
        room_num = r[1]
        room_type = r[2]
        sno = r[3]
        name = r[4]
        committee = r[5]
        post = r[6]
        
        if floor: current_floor = str(floor).strip()
        if room_num: current_room = str(room_num).strip()
        if room_type: current_type = str(room_type).strip()
        
        if not current_room or sno is None:
            continue
            
        try:
            sno_val = int(sno)
        except ValueError:
            continue
            
        # Determine bed representation, e.g. "A-1", "A-2", "B", "C-1" etc.
        # We can just count the index of this bed within the current room & type
        # Or simply label it "Type A Bed 1" or similar. Let's make it clean: "Bed A-1"
        rooms.append({
            'hostel': 'GH-2',
            'floor': current_floor,
            'room': current_room,
            'bed': f"Bed {current_type}" if current_type else "Bed",
            'sno': sno_val,
            'allottedTo': None,
            'allottedToName': None,
            'allottedToAppNo': None
        })
        
    # Post-process bed labels to distinguish double beds of same type (e.g. A-1 and A-2)
    room_type_counts = {}
    for r in rooms:
        key = (r['room'], r['bed'])
        room_type_counts[key] = room_type_counts.get(key, 0) + 1
        
    # Second pass: append bed index if there are multiples of the same type in that room
    seen_counts = {}
    for r in rooms:
        key = (r['room'], r['bed'])
        total = room_type_counts[key]
        if total > 1:
            seen_counts[key] = seen_counts.get(key, 0) + 1
            # E.g. "Bed A" -> "Bed A-1"
            r['bed'] = f"{r['bed']}-{seen_counts[key]}"
            
    return rooms

def run():
    uri = get_mongo_uri()
    if not uri:
        print("Error: MONGODB_URI not found.")
        return
        
    client = MongoClient(uri)
    db = client['test']
    hostel_rooms_coll = db['hostelrooms']
    
    # Load Excel
    print(f"Loading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Parse sheets
    bh1_rooms = parse_bh1(wb['BH-1'])
    gh2_rooms = parse_gh2(wb['GH-2'])
    
    all_rooms = bh1_rooms + gh2_rooms
    print(f"Parsed {len(bh1_rooms)} BH-1 beds and {len(gh2_rooms)} GH-2 beds. Total = {len(all_rooms)}")
    
    # Clear and insert into DB
    print("Clearing existing hostel rooms in MongoDB...")
    hostel_rooms_coll.delete_many({})
    
    print("Inserting rooms into MongoDB...")
    if all_rooms:
        hostel_rooms_coll.insert_many(all_rooms)
        
    print("Successfully seeded hostel rooms database!")

if __name__ == '__main__':
    run()
